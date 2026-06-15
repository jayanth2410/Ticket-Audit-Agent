"""
Excel Handler — audit report generation
========================================
Holds the workbook in memory and writes all rows before a single save().
"""

import shutil
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from typing import Dict, Any

# ── Column → metric key + failure label ───────────────────────────────────────
METRIC_MAP = {
    "F": ("response_within_sla",       "Response SLA not met"),
    "G": ("short_desc_quality",        "Short description unclear"),
    "H": ("priority_reassessed",       "Priority not re-assessed"),
    "I": ("incident_reassigned",       "Reassignment details missing"),
    "J": ("user_contact",              "User contact not documented"),
    "K": ("pending_status",            "Pending status incorrectly used"),
    "L": ("work_notes_regular_update", "Work notes not updated regularly"),
    "M": ("resolution_notes_quality",  "Resolution notes incomplete"),
    "N": ("resolution_sla",            "Resolution SLA not met"),
    "O": ("user_confirmation",         "User confirmation not taken"),
    "P": ("reopened_user_connect",     "No user contact after reopen"),
    "Q": ("kba_education",             "KBA not shared with user"),
}

# Max points per metric column
METRIC_MAX_SCORES = {
    "F": 5,
    "G": 5,
    "H": 10,
    "I": 10,
    "J": 10,
    "K": 5,
    "L": 15,
    "M": 15,
    "N": 10,
    "O": 5,
    "P": 5,
    "Q": 5,
}

# ── Styles ────────────────────────────────────────────────────────────────────
PASS_FILL = PatternFill("solid", start_color="00C851", end_color="00C851")
FAIL_FILL = PatternFill("solid", start_color="FF4444", end_color="FF4444")
PASS_FONT = Font(bold=True, color="FFFFFF")
FAIL_FONT = Font(bold=True, color="FFFFFF")
CENTER    = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT      = Alignment(horizontal="left",   vertical="center", wrap_text=True)


class ExcelHandler:
    """
    Duplicate the template once, then keep the workbook open in memory.
    Call write_audit_row() for each ticket, then save() once at the end.
    """

    def __init__(self, template_path: str, output_path: str, pass_threshold: float = 70):
        self.template_path  = template_path
        self.output_path    = output_path
        self.pass_threshold = pass_threshold
        self.current_row    = 3   # data rows start at row 3

        # Copy template → output
        shutil.copy(template_path, output_path)
        print(f"Template copied → {output_path}")

        # Load workbook once and keep it open
        self._wb = load_workbook(output_path)
        self._ws = self._wb.active

        # One-time setup on the workbook
        self._write_max_score_row()
        self._add_dropdown_validation()

    # ── One-time setup ────────────────────────────────────────────────────────

    def _write_max_score_row(self):
        """Write max-point values into row 2 for each metric column."""
        for col, pts in METRIC_MAX_SCORES.items():
            cell           = self._ws[f"{col}2"]
            cell.value     = pts
            cell.alignment = CENTER

    def _add_dropdown_validation(self):
        """Add Yes / No / NA dropdown to metric columns F:Q for all data rows."""
        dv = DataValidation(
            type         = "list",
            formula1     = '"Yes,No,NA"',
            allow_blank  = True,
            showDropDown = False,
        )
        dv.sqref = "F3:Q2000"
        self._ws.add_data_validation(dv)

    # ── Score helpers ─────────────────────────────────────────────────────────

    def _compute_score(self, auditor_data: Dict[str, Any]) -> Dict[str, Any]:
        """NA metrics are excluded from both numerator and denominator."""
        score  = 0
        out_of = 0

        for col, (key, _) in METRIC_MAP.items():
            value   = auditor_data.get(key, "NA")
            max_pts = METRIC_MAX_SCORES[col]
            if value == "NA":
                continue
            out_of += max_pts
            if value == "Yes":
                score  += max_pts

        percentage     = round(score / out_of * 100, 1) if out_of > 0 else 0.0
        quality_result = "PASS" if percentage >= self.pass_threshold else "FAIL"

        return {
            "score"         : score,
            "out_of"        : out_of,
            "percentage"    : percentage,
            "quality_result": quality_result,
        }

    def _build_observation(self, auditor_data: Dict[str, Any]) -> str:
        failed = [
            label
            for _, (key, label) in METRIC_MAP.items()
            if auditor_data.get(key, "NA") == "No"
        ]
        return ("; ".join(failed) + ".") if failed else "All applicable metrics passed."

    # ── Public write method ───────────────────────────────────────────────────

    def write_audit_row(self, auditor_data: Dict[str, Any]) -> None:
        """Write one audit row to the in-memory workbook (does not save to disk)."""
        row    = self.current_row
        ws     = self._ws
        scores = self._compute_score(auditor_data)

        # ── A–E  header info ──────────────────────────────────────────────────
        for col, key in zip(
            ["A", "B", "C", "D", "E"],
            ["ticket_number", "created_by", "priority", "tcs_resolver_group", "resolved_by"],
        ):
            ws[f"{col}{row}"] = auditor_data.get(key, "")
            ws[f"{col}{row}"].alignment = LEFT

        # ── F–Q  metric values ────────────────────────────────────────────────
        for col, (key, _) in METRIC_MAP.items():
            cell           = ws[f"{col}{row}"]
            cell.value     = auditor_data.get(key, "NA")
            cell.alignment = CENTER

        # ── R  Score ──────────────────────────────────────────────────────────
        ws[f"R{row}"]           = scores["score"]
        ws[f"R{row}"].alignment = CENTER

        # ── S  Out of ─────────────────────────────────────────────────────────
        ws[f"S{row}"]           = scores["out_of"]
        ws[f"S{row}"].alignment = CENTER

        # ── T  Percentage ─────────────────────────────────────────────────────
        ws[f"T{row}"]           = f"{scores['percentage']}%"
        ws[f"T{row}"].alignment = CENTER

        # ── U  Quality result ─────────────────────────────────────────────────
        ws[f"U{row}"]           = scores["quality_result"]
        ws[f"U{row}"].alignment = CENTER

        if scores["quality_result"] == "PASS":
            ws[f"U{row}"].fill = PASS_FILL
            ws[f"U{row}"].font = PASS_FONT
        else:
            ws[f"U{row}"].fill = FAIL_FILL
            ws[f"U{row}"].font = FAIL_FONT

        # ── V  Observation ────────────────────────────────────────────────────
        ws[f"V{row}"]           = self._build_observation(auditor_data)
        ws[f"V{row}"].alignment = LEFT

        print(
            f"  Row {row}: {auditor_data.get('ticket_number', 'N/A')} "
            f"— {scores['score']}/{scores['out_of']} "
            f"({scores['percentage']}%) {scores['quality_result']}"
        )

        self.current_row += 1

    def save(self):
        """Persist the workbook to disk. Call once after all rows are written."""
        self._wb.save(self.output_path)
        print(f"Excel report saved → {self.output_path}")

    def get_current_row(self) -> int:
        return self.current_row
